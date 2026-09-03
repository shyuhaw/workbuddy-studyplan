# -*- coding: utf-8 -*-
"""
Excel 快速探查器
用法: python excel_probe.py <文件路径>
输出: 所有 sheet 的行列数 + 每个 sheet 的表头 + 第一个 sheet 前 3 行预览

支持 .xlsx (openpyxl) 和 .xls (xlrd)
"""
import sys
import os

# 依赖装在隔离目录，脚本自包含
_PKGS = r"C:\Users\Administrator\.workbuddy\binaries\python\pkgs"
if _PKGS not in sys.path:
    sys.path.insert(0, _PKGS)

import openpyxl
import xlrd


def probe_xlsx(path):
    """读取 xlsx，返回 sheet 信息列表"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    info = []
    for name in wb.sheetnames:
        ws = wb[name]
        info.append({
            "name": name,
            "rows": ws.max_row,
            "cols": ws.max_column,
        })
    wb.close()
    return info


def probe_xls(path):
    """读取 xls，返回 sheet 信息列表"""
    wb = xlrd.open_workbook(path)
    info = []
    for ws in wb.sheets():
        info.append({
            "name": ws.name,
            "rows": ws.nrows,
            "cols": ws.ncols,
        })
    return info


def preview(path, ext, sheet_idx=0, n=3):
    """预览指定 sheet 的前 n 行"""
    if ext == ".xlsx":
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[sheet_idx]]
        out = []
        for i, row in enumerate(ws.iter_rows(max_row=n, values_only=True)):
            out.append([("" if c is None else str(c))[:20] for c in row])
        wb.close()
        return out
    else:
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(sheet_idx)
        out = []
        for i in range(min(n, ws.nrows)):
            out.append([str(c)[:20] for c in ws.row_values(i)])
        return out


def main():
    if len(sys.argv) < 2:
        print("用法: python excel_probe.py <Excel文件路径>")
        return 1

    path = sys.argv[1]
    if not os.path.exists(path):
        print(f"[!] 文件不存在: {path}")
        return 1

    ext = os.path.splitext(path)[1].lower()
    size_kb = os.path.getsize(path) / 1024

    print(f"[文件] {os.path.basename(path)}")
    print(f"[大小] {size_kb:.1f} KB")
    print(f"[格式] {ext}")
    print("=" * 55)

    if ext == ".xlsx":
        info = probe_xlsx(path)
    elif ext == ".xls":
        info = probe_xls(path)
    else:
        print(f"[!] 不支持的格式: {ext}")
        return 1

    print(f"[Sheet 数量] {len(info)}")
    print("-" * 55)
    for i, s in enumerate(info):
        mark = " ← 第一个" if i == 0 else ""
        print(f"  {i + 1}. {s['name']}  |  {s['rows']} 行 x {s['cols']} 列{mark}")

    print("=" * 55)
    print(f"【前 3 行预览】(Sheet: {info[0]['name']})")
    print("-" * 55)
    for i, row in enumerate(preview(path, ext), 1):
        print(f"  第{i}行: {row}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
