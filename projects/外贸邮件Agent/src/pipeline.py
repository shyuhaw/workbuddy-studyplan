# -*- coding: utf-8 -*-
"""
邮件分类流水线 v1.1（规则层 + LLM 兜底）
==========================================
架构：
    邮件 → [规则层 classifier.py] → 置信度判定
                                      ├─ ≥阈值 → 直接采纳（零成本，毫秒级）
                                      └─ <阈值 → [LLM兜底层 llm_fallback.py] → 最终分类

设计要点：
1. 规则层负责"快"和"省"——处理 80% 的明确邮件
2. LLM 层负责"准"——只处理规则层拿不准的那 20%
3. 人工只需复核最难的 5%——成本、准确率、人力三者平衡

用法：
    python pipeline.py [数据文件] [--threshold 0.8] [--no-llm]

作者：麦当
日期：2026-08-31
"""

import os
import sys
import json
import argparse
from collections import Counter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

_PKGS = r"C:\Users\Administrator\.workbuddy\binaries\python\pkgs"
if _PKGS not in sys.path:
    sys.path.insert(0, _PKGS)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from classifier import score_email, LABEL_CN
from llm_fallback import FallbackManager

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_DATA = os.path.join(BASE_DIR, "data", "ambiguous_emails.json")
DEFAULT_OUT = os.path.join(BASE_DIR, "output", "流水线结果.xlsx")


def process(emails, threshold=0.8, use_llm=True):
    """执行完整流水线"""
    fm = FallbackManager(threshold=threshold) if use_llm else None
    results = []

    for mail in emails:
        subject = mail.get("subject", "")
        body = mail.get("body", "")
        expected = mail.get("expected", "")

        # ---- 第一层：规则层 ----
        rule_cat, rule_conf, scores, hits = score_email(subject, body)

        # ---- 第二层：置信度判定 ----
        need_fb = use_llm and fm.need_fallback(rule_conf)

        final_cat = rule_cat
        final_reason = "规则层直接判定（置信度达标）"
        llm_cat = ""
        llm_reason = ""
        llm_conf = 0.0

        if need_fb:
            llm_cat, llm_reason, llm_conf = fm.classify(subject, body, scores)
            final_cat = llm_cat
            final_reason = llm_reason

        results.append({
            "id": mail["id"],
            "from": mail.get("from", ""),
            "subject": subject,
            "rule_cat": LABEL_CN.get(rule_cat, rule_cat),
            "rule_conf": round(rule_conf, 3),
            "need_fallback": need_fb,
            "llm_cat": LABEL_CN.get(llm_cat, llm_cat) if llm_cat else "",
            "llm_reason": llm_reason,
            "llm_conf": round(llm_conf, 2),
            "final_cat": LABEL_CN.get(final_cat, final_cat),
            "expected": LABEL_CN.get(expected, expected),
            "rule_correct": rule_cat == expected,
            "final_correct": final_cat == expected,
            "hits": hits[:4],
            "note": mail.get("note", ""),
        })

    return results, fm


def export_excel(results, path, meta):
    """导出对比报告"""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "流水线结果"

    headers = [
        "邮件ID", "主题", "规则层判定", "规则层置信度", "是否触发兜底",
        "LLM判定", "LLM理由", "最终判定", "正确答案", "规则层是否正确", "最终是否正确"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wrong_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    right_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fb_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for r in results:
        ws.append([
            r["id"],
            r["subject"][:45],
            r["rule_cat"],
            f"{r['rule_conf']:.1%}",
            "是" if r["need_fallback"] else "否",
            r["llm_cat"] or "-",
            r["llm_reason"][:40],
            r["final_cat"],
            r["expected"],
            "✓" if r["rule_correct"] else "✗",
            "✓" if r["final_correct"] else "✗",
        ])

        row_idx = ws.max_row
        # 触发兜底的行标黄
        if r["need_fallback"]:
            for cell in ws[row_idx]:
                cell.fill = fb_fill
        # 规则层错但兜底救回来的，最终列标绿
        if not r["rule_correct"] and r["final_correct"]:
            ws.cell(row=row_idx, column=11).fill = right_fill
            ws.cell(row=row_idx, column=11).font = Font(bold=True, color="006100")
        # 最终仍错的标红
        if not r["final_correct"]:
            ws.cell(row=row_idx, column=11).fill = wrong_fill
            ws.cell(row=row_idx, column=11).font = Font(bold=True, color="9C0006")

    widths = [8, 42, 12, 13, 12, 11, 38, 11, 11, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 汇总
    ws.append([])
    ws.append(["【汇总】"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    total = len(results)
    rule_ok = sum(1 for r in results if r["rule_correct"])
    final_ok = sum(1 for r in results if r["final_correct"])
    fb_count = sum(1 for r in results if r["need_fallback"])

    for label, value in [
        ("样本总数", total),
        ("规则层正确", f"{rule_ok}/{total} = {rule_ok/total:.1%}"),
        ("兜底后正确", f"{final_ok}/{total} = {final_ok/total:.1%}"),
        ("触发兜底数量", f"{fb_count}/{total} = {fb_count/total:.1%}"),
        ("兜底挽回", f"{final_ok - rule_ok} 封"),
        ("LLM Provider", f"{meta['provider']}{'（模拟）' if meta['is_mock'] else ''}"),
        ("LLM 调用次数", meta["calls"]),
    ]:
        ws.append([label, value])

    wb.save(path)
    return path


def main():
    ap = argparse.ArgumentParser(description="邮件分类流水线 v1.1")
    ap.add_argument("data", nargs="?", default=DEFAULT_DATA, help="数据文件路径")
    ap.add_argument("--threshold", type=float, default=0.8, help="兜底阈值，默认 0.8")
    ap.add_argument("--no-llm", action="store_true", help="只跑规则层，用于对比")
    ap.add_argument("--out", default=DEFAULT_OUT, help="输出 Excel 路径")
    args = ap.parse_args()

    with open(args.data, "r", encoding="utf-8") as f:
        emails = json.load(f)

    print("=" * 78)
    print("邮件分类流水线 v1.1  ——  规则层 + LLM 兜底")
    print("=" * 78)
    print(f"数据文件: {os.path.basename(args.data)}")
    print(f"样本数量: {len(emails)}")
    print(f"兜底阈值: {args.threshold:.0%}")
    print(f"LLM 兜底: {'关闭（仅规则层）' if args.no_llm else '开启'}")
    print()

    results, fm = process(emails, threshold=args.threshold, use_llm=not args.no_llm)
    meta = fm.summary() if fm else {"provider": "none", "is_mock": True, "calls": 0}

    if fm:
        tag = "（模拟模式，非真实 LLM）" if meta["is_mock"] else "（真实 API）"
        print(f"LLM Provider: {meta['provider']} {tag}")
        print()

    # 明细
    print(f"{'ID':<5}{'规则层':<8}{'置信':<8}{'兜底':<6}{'LLM':<8}{'最终':<8}{'正确':<6}")
    print("-" * 78)
    for r in results:
        fb = "是" if r["need_fallback"] else "-"
        llm = r["llm_cat"] or "-"
        mark = "✓" if r["final_correct"] else "✗"
        # 标记兜底挽回的
        if not r["rule_correct"] and r["final_correct"]:
            mark = "✓ 挽回"
        print(f"{r['id']:<5}{r['rule_cat']:<8}{r['rule_conf']:<8.1%}{fb:<6}"
              f"{llm:<8}{r['final_cat']:<8}{mark:<6}")

    # 统计
    total = len(results)
    rule_ok = sum(1 for r in results if r["rule_correct"])
    final_ok = sum(1 for r in results if r["final_correct"])
    fb_count = sum(1 for r in results if r["need_fallback"])

    print("-" * 78)
    print(f"规则层准确率:   {rule_ok}/{total} = {rule_ok/total:.1%}")
    print(f"兜底后准确率:   {final_ok}/{total} = {final_ok/total:.1%}")
    print(f"触发兜底:       {fb_count}/{total} = {fb_count/total:.1%}")
    print(f"兜底挽回:       {final_ok - rule_ok} 封")
    print()

    # 展示兜底理由
    fb_results = [r for r in results if r["need_fallback"]]
    if fb_results:
        print("【LLM 兜底详情】")
        print("-" * 78)
        for r in fb_results:
            print(f"\n[{r['id']}] {r['subject'][:50]}")
            print(f"  规则层: {r['rule_cat']} ({r['rule_conf']:.1%})  →  LLM: {r['llm_cat']}")
            print(f"  理由: {r['llm_reason']}")
            print(f"  正确答案: {r['expected']}  {'✓ 对' if r['final_correct'] else '✗ 错'}")
        print()

    out = export_excel(results, args.out, meta)
    print("=" * 78)
    print(f"[已导出] {out}")
    print("=" * 78)

    return 0


if __name__ == "__main__":
    sys.exit(main())
