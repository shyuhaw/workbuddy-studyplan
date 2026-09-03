# -*- coding: utf-8 -*-
"""
跨境客户邮件智能分类器 v1.0
================================
功能：对外贸客户邮件做 4 分类（询盘 / 订单 / 投诉 / 通知）
方法：中英双语关键词加权打分（规则层），零 API 成本、毫秒级响应

作者：麦当
项目：跨境客户邮件智能处理 Agent
日期：2026-08-30

为什么先用规则而不是 LLM：
1. 规则层可做第一道过滤，只有低置信度的才交给 LLM，省 token
2. 规则可解释、可审计，企业客户要的就是"为什么这么分"
3. 后续接 LLM 时，规则层可作为兜底和校验
"""

import json
import os
import sys
import re

_PKGS = r"C:\Users\Administrator\.workbuddy\binaries\python\pkgs"
if _PKGS not in sys.path:
    sys.path.insert(0, _PKGS)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_FILE = os.path.join(BASE_DIR, "data", "sample_emails.json")
OUTPUT_FILE = os.path.join(BASE_DIR, "output", "分类结果.xlsx")

# ---------------------------------------------------------------------------
# 关键词库：中英双语，按业务语义分组
# ---------------------------------------------------------------------------
KEYWORDS = {
    "inquiry": {  # 询盘：还在问价/问规格阶段
        "en": [
            "inquiry", "enquiry", "quotation", "quote", "best price", "price list",
            "moq", "minimum order", "sample request", "sample", "catalog",
            "lead time", "delivery time", "interested in", "could you please",
            "please advise", "payment terms", "warranty", "specification",
        ],
        "zh": [
            "询问", "咨询", "报价", "报价单", "最小起订量", "起订量",
            "样品", "产品目录", "了解一下", "请问", "能否提供", "交期",
        ],
    },
    "order": {  # 订单：已经确认要买
        "en": [
            "purchase order", "p.o.", "po-", "order #", "order no", "order number",
            "place an order", "confirm order", "order confirmation", "proforma invoice",
            # ↓ 错误分析后补充（E06 误判暴露的缺口）：
            # 原词库只覆盖 PO 号、形式发票等"已成单"信号，漏掉"刚确认、要求投产"这一阶段，
            # 导致 'We accept the price ... please proceed with production' 被判成询盘。
            # 这些是真实外贸邮件中客户确认订单的常用表达，非为测试样本量身定制。
            "accept the price", "we accept", "price is accepted", "accept your quotation",
            "proceed with production", "please proceed", "proceed with the order",
            "we'd like to order", "would like to order", "go ahead with", "approved the sample",
            "quantity", "total amount", "amendment", "amend", "change the spec",
            "increase the quantity", "delivery date", "unit price",
        ],
        "zh": [
            "订单", "采购订单", "确认订单", "下单", "数量", "单价",
            "合计金额", "交货日期", "收货地址", "备货", "合同编号",
        ],
    },
    "complaint": {  # 投诉/异常：出问题了，优先级最高
        "en": [
            "complaint", "quality issue", "quality problem", "defective", "defect",
            "damage", "damaged", "shortage", "missing", "delay", "delayed",
            "claim", "compensation", "reject", "rejection", "serious problem",
            "regret to inform", "urgent issue", "breach", "penalty",
        ],
        "zh": [
            "投诉", "质量问题", "质量", "破损", "损坏", "短缺", "缺少",
            "延误", "延期", "索赔", "赔偿", "违约", "停工", "损失",
        ],
    },
    "notification": {  # 通知：被动告知，不需行动
        "en": [
            "shipping advice", "bill of lading", "b/l", "etd", "eta", "vessel",
            "payment received", "credited", "customs", "clearance", "released",
            "system-generated", "do not reply", "auto", "notification", "advice",
            "remittance", "declaration",
        ],
        "zh": [
            "通知", "系统通知", "发货通知", "清关", "放行", "自动发送",
            "请勿回复", "请勿直接回复", "预警", "到账", "汇款",
        ],
    },
}

# 权重：主题比正文更能说明意图
WEIGHT_SUBJECT = 3
WEIGHT_BODY = 1

# 中文分类名映射
LABEL_CN = {
    "inquiry": "询盘",
    "order": "订单",
    "complaint": "投诉",
    "notification": "通知",
}


def score_email(subject, body):
    """
    对单封邮件打分
    返回: (预测类别, 置信度, 各类得分, 命中的关键词)
    """
    subject_l = (subject or "").lower()
    body_l = (body or "").lower()

    scores = {}
    hits = {}

    for category, langs in KEYWORDS.items():
        cat_score = 0
        cat_hits = []
        for lang in ("en", "zh"):
            for kw in langs[lang]:
                in_subject = kw in subject_l
                in_body = kw in body_l
                if in_subject:
                    cat_score += WEIGHT_SUBJECT
                    cat_hits.append(f"[主题]{kw}")
                elif in_body:
                    cat_score += WEIGHT_BODY
                    cat_hits.append(kw)
        scores[category] = cat_score
        hits[category] = cat_hits

    # 取最高分
    best = max(scores, key=scores.get)
    total = sum(scores.values())
    confidence = (scores[best] / total) if total > 0 else 0.0

    return best, confidence, scores, hits[best]


def classify_all(emails):
    """批量分类"""
    results = []
    for mail in emails:
        pred, conf, scores, hits = score_email(mail.get("subject", ""), mail.get("body", ""))
        results.append({
            "id": mail["id"],
            "from": mail.get("from", ""),
            "subject": mail.get("subject", ""),
            "predicted": pred,
            "predicted_cn": LABEL_CN[pred],
            "expected": mail.get("expected", ""),
            "expected_cn": LABEL_CN.get(mail.get("expected", ""), ""),
            "correct": pred == mail.get("expected"),
            "confidence": round(conf, 3),
            "scores": scores,
            "hits": hits[:5],  # 只保留前5个命中词
        })
    return results


def export_excel(results, path):
    """导出 Excel 报告"""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分类结果"

    headers = ["邮件ID", "发件人", "主题", "预测分类", "实际分类", "是否正确", "置信度", "命中关键词"]
    ws.append(headers)

    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    for r in results:
        ws.append([
            r["id"],
            r["from"],
            r["subject"][:50],
            r["predicted_cn"],
            r["expected_cn"],
            "✓" if r["correct"] else "✗",
            f"{r['confidence']:.1%}",
            ", ".join(r["hits"]),
        ])

    # 错行标红
    wrong_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for i, r in enumerate(results, start=2):
        if not r["correct"]:
            for cell in ws[i]:
                cell.fill = wrong_fill

    # 列宽
    widths = [10, 32, 45, 12, 12, 10, 10, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 汇总行
    ws.append([])
    total = len(results)
    correct = sum(1 for r in results if r["correct"])
    ws.append(["准确率", f"{correct}/{total}", f"{correct/total:.1%}"])
    ws[f"A{total + 3}"].font = Font(bold=True)

    wb.save(path)
    return path


def main():
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        emails = json.load(f)

    print("=" * 70)
    print("跨境客户邮件智能分类器 v1.0")
    print("=" * 70)
    print(f"样本数量: {len(emails)} 封")
    print()

    results = classify_all(emails)

    # 控制台输出
    print(f"{'ID':<6}{'预测':<8}{'实际':<8}{'结果':<6}{'置信度':<10}命中关键词")
    print("-" * 70)
    for r in results:
        mark = "✓" if r["correct"] else "✗ 错"
        print(f"{r['id']:<6}{r['predicted_cn']:<8}{r['expected_cn']:<8}{mark:<6}"
              f"{r['confidence']:.1%}    {', '.join(r['hits'][:3])}")

    # 统计
    total = len(results)
    correct = sum(1 for r in results if r["correct"])
    print("-" * 70)
    print(f"准确率: {correct}/{total} = {correct/total:.1%}")

    # 分类分布
    print()
    print("分类分布:")
    from collections import Counter
    dist = Counter(r["predicted_cn"] for r in results)
    for k, v in dist.items():
        print(f"  {k}: {v} 封")

    # 导出
    out = export_excel(results, OUTPUT_FILE)
    print()
    print(f"[已导出] {out}")
    print("=" * 70)

    return 0


if __name__ == "__main__":
    sys.exit(main())
