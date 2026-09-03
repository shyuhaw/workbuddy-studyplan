# -*- coding: utf-8 -*-
"""
邮件接入层 v1.0 —— 把真实邮件喂给端到端 Agent
==========================================
分工：
    取邮件 —— WorkBuddy 的 Agent Mail 连接器（MCP）负责
    处理   —— 本脚本负责（本地 Python）

数据流：
    Agent Mail 收件箱
        ↓  MCP 拉取，落成 data/inbox_snapshot.json
    mail_connector.py
        ↓
    MailAgent（分类 → 提取 → 决策）
        ↓
    output/真实邮件处理结果.xlsx

用法：
    python mail_connector.py [--input data/inbox_snapshot.json] [--limit N]

⚠️ 隐私边界（真实使用时必读）：
    邮件正文会发送给 LLM（DeepSeek）做分类与提取，即客户名、产品、价格会离开本机。
    - 含底价、合同条款、个人身份信息的邮件 → 先脱敏，或加 --no-llm 只跑本地规则层
    - 本脚本不落盘原始正文到任何对外位置，快照文件仅存本机

作者：麦当
日期：2026-08-31
"""

import os
import sys
import json
import argparse
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

_PKGS = r"C:\Users\Administrator\.workbuddy\binaries\python\pkgs"
if _PKGS not in sys.path:
    sys.path.insert(0, _PKGS)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from agent import MailAgent, parse_deadline_days, FIELD_CN, BASE_DATE
from classifier import LABEL_CN
from extractor import FIELDS

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_IN = os.path.join(BASE_DIR, "data", "inbox_snapshot.json")
DEFAULT_OUT = os.path.join(BASE_DIR, "output", "真实邮件处理结果.xlsx")


# ---------------------------------------------------------------------------
# 正文清洗 —— 真实邮件远比本地 JSON 脏
# ---------------------------------------------------------------------------
import re
import html as html_lib

# 历史引用分隔（旧报价/旧数量会污染提取 —— E06 已验证这类坑）
QUOTE_MARKERS = [
    r"-{3,}\s*(?:原始邮件|Original Message|原始消息|Original message)\s*-{3,}",
    r"_{10,}",
    r"On\s+.{5,80}wrote:",
    r"在\s*.{5,60}\s*写道[：:]",
    r"^\s*发件人\s*[:：]",
]
# 服务商页脚 / 免责声明（含 token 与退订链接，数字污染数量、域名污染客户）
FOOTER_MARKERS = [
    "此邮件由", "本邮件由", "通过 Agent Mail 自动发送",
    "举报", "退订", "请勿直接回复本邮件",
    "This email and any attachments are confidential",
    "CONFIDENTIALITY NOTICE", "免责声明",
]


def html_to_text(raw):
    """HTML → 纯文本，保留换行结构"""
    t = re.sub(r"<(script|style)[^>]*>.*?</\1>", "", raw, flags=re.S | re.I)
    t = re.sub(r"<br\s*/?\s*>", "\n", t, flags=re.I)
    t = re.sub(r"</(p|div|tr|li|h\d)\s*>", "\n", t, flags=re.I)
    t = re.sub(r"<[^>]+>", "", t)
    t = html_lib.unescape(t)
    t = re.sub(r"[ \t\u00a0]+", " ", t)
    t = re.sub(r"\n\s*\n\s*\n+", "\n\n", t)
    return t.strip()


def strip_noise(text):
    """砍掉历史引用与页脚两类噪音"""
    for pat in QUOTE_MARKERS:
        m = re.search(pat, text, re.I | re.M)
        if m and m.start() > len(text) * 0.15:   # 避免误伤邮件开头
            text = text[:m.start()]

    cut = len(text)
    for marker in FOOTER_MARKERS:
        idx = text.find(marker)
        if idx > 0:
            cut = min(cut, idx)
    return text[:cut].strip()


def clean_body(body):
    """HTML/纯文本 → 干净正文。这是接真实邮箱后才暴露出来的必要一步。"""
    if not body:
        return ""
    text = html_to_text(body) if ("<" in body and ">" in body) else body
    return strip_noise(text)


# ---------------------------------------------------------------------------
# 加载快照
# ---------------------------------------------------------------------------
def load_snapshot(path, limit=None):
    """加载 Agent Mail 拉取下来的邮件快照（自动清洗正文）"""
    with open(path, "r", encoding="utf-8") as f:
        mails = json.load(f)
    if limit:
        mails = mails[:limit]
    # 统一字段，兼容不同来源
    out = []
    for i, m in enumerate(mails, 1):
        raw_body = m.get("body", "") or m.get("snippet", "")
        clean = clean_body(raw_body)
        out.append({
            "id": m.get("id") or m.get("message_id") or f"M{i:03d}",
            "from": m.get("from", ""),
            "subject": m.get("subject", ""),
            "body": clean,
            "raw_len": len(raw_body),
            "clean_len": len(clean),
            "date": m.get("date", "") or m.get("created_at", ""),
        })
    return out


# ---------------------------------------------------------------------------
# 输出（真实邮件无标注，故不统计准确率，改统计"字段完整度"）
# ---------------------------------------------------------------------------
def export(results, path, meta):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "处理结果"
    headers = ["#", "收到时间", "发件人", "主题", "分类", "分类来源", "客户", "产品",
               "数量", "价格", "截止日", "剩余天数", "优先级", "判断依据", "建议动作", "风险提示"]
    ws.append(headers)
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in ws[1]:
        c.fill = hf
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    high = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    mid = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    low = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for i, r in enumerate(results, 1):
        f = r["fields"]
        days = parse_deadline_days(f.get("deadline"))
        ws.append([
            i, r.get("date", "")[:19].replace("T", " "),
            r.get("from", ""), r["subject"][:50],
            LABEL_CN.get(r["category"], r["category"]), r["cat_source"],
            f.get("customer") or "—", f.get("product") or "—",
            f.get("quantity") or "—", f.get("price") or "—",
            f.get("deadline") or "—",
            f"{days} 天" if days is not None else "—",
            r["priority"],
            "；".join(r["reasons"]),
            "；".join(r["actions"]),
            "；".join(r["risks"]) or "无",
        ])
        cell = ws.cell(row=ws.max_row, column=13)
        cell.fill = {"高": high, "中": mid}.get(r["priority"], low)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for i, w in enumerate([5, 19, 24, 42, 8, 11, 22, 30, 22, 20, 20, 10, 8, 32, 40, 34], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---------- 汇总 ----------
    n = len(results)
    total_f = n * len(FIELDS)
    filled = sum(1 for r in results for f in FIELDS if r["fields"].get(f))
    llm_fields = sum(len(r["ext_targets"]) for r in results)

    ws2 = wb.create_sheet("运行统计")
    ws2.append(["【真实邮件处理 · 运行统计】"])
    ws2.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws2.append([])
    for k, v in [
        ("处理时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("邮件数量", n),
        ("字段完整度", f"{filled}/{total_f} = {filled/total_f:.1%}（成功抽出的字段占比）"),
        ("分类触发 LLM", f"{sum(1 for r in results if r['need_cls'])} 封"),
        ("提取触发 LLM", f"{llm_fields} 个字段 / "
                        f"{sum(1 for r in results if r['ext_targets'])} 封"),
        ("LLM 调用合计", f"{meta['cls_calls'] + meta['ext_calls']} 次"
                        f"（分类 {meta['cls_calls']} + 提取 {meta['ext_calls']}）"),
        ("Provider", f"{meta['provider']}{'（模拟）' if meta['is_mock'] else '（真实）'}"),
        ("优先级分布", "  ".join(f"{p} {sum(1 for r in results if r['priority'] == p)} 封"
                               for p in ("高", "中", "低"))),
        ("基准日期", str(BASE_DATE)),
    ]:
        ws2.append([k, v])

    # 分类分布
    ws2.append([])
    ws2.append(["分类分布", ""])
    dist = {}
    for r in results:
        k = LABEL_CN.get(r["category"], r["category"])
        dist[k] = dist.get(k, 0) + 1
    for k, v in dist.items():
        ws2.append([f"  {k}", f"{v} 封"])

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 46

    wb.save(path)
    return {"n": n, "filled": filled, "total_f": total_f,
            "llm_fields": llm_fields, "path": path}


def main():
    ap = argparse.ArgumentParser(description="邮件接入层 v1.0")
    ap.add_argument("--input", default=DEFAULT_IN, help="Agent Mail 快照 JSON")
    ap.add_argument("--limit", type=int, default=None, help="只处理前 N 封")
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()

    if not os.path.exists(args.input):
        print(f"[错误] 找不到快照文件: {args.input}")
        print("请先用 Agent Mail 连接器拉取收件箱，导出为该 JSON。")
        return 1

    mails = load_snapshot(args.input, args.limit)

    print("=" * 90)
    print("邮件接入层 v1.0  ——  真实邮件 → Agent → 结果")
    print("=" * 90)
    print(f"快照文件: {os.path.basename(args.input)}")
    print(f"邮件数量: {len(mails)}")
    print(f"基准日期: {BASE_DATE}\n")

    agent = MailAgent()
    print(f"LLM Provider: {agent.em.provider.name} "
          f"{'（模拟）' if agent.em.is_mock else '（真实）'}\n")

    results = agent.process(mails)
    for r in results:
        r["from"] = next((m["from"] for m in mails if m["id"] == r["id"]), "")
        r["date"] = next((m.get("date", "") for m in mails if m["id"] == r["id"]), "")

    print(f"{'#':<4}{'分类':<7}{'来源':<9}{'优先级':<7}{'客户':<20}{'数量':<20}{'截止日':<18}")
    print("-" * 90)
    for i, r in enumerate(results, 1):
        f = r["fields"]
        print(f"{i:<4}{LABEL_CN.get(r['category'], r['category']):<7}{r['cat_source']:<9}"
              f"{r['priority']:<7}{str(f.get('customer') or '—')[:18]:<20}"
              f"{str(f.get('quantity') or '—')[:18]:<20}"
              f"{str(f.get('deadline') or '—')[:16]:<18}")

    meta = {
        "provider": agent.em.provider.name,
        "is_mock": agent.em.is_mock,
        "cls_calls": agent.fm.call_count,
        "ext_calls": agent.em.call_count,
    }
    stat = export(results, args.out, meta)

    print("-" * 90)
    print(f"字段完整度: {stat['filled']}/{stat['total_f']} = {stat['filled']/stat['total_f']:.1%}")
    print(f"LLM 调用:   分类 {meta['cls_calls']} 次 + 提取 {meta['ext_calls']} 次"
          f" = {meta['cls_calls'] + meta['ext_calls']} 次")
    print("=" * 90)
    print(f"[已导出] {stat['path']}")
    print("=" * 90)
    return 0


if __name__ == "__main__":
    sys.exit(main())
