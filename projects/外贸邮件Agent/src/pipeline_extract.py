# -*- coding: utf-8 -*-
"""
信息提取流水线 v1.0（规则层 + 字段级 LLM 兜底）
==========================================
架构：
    邮件 → [规则层 extractor.py] → 逐字段置信度判定
                                     ├─ 高置信(≥阈值) → 直接采纳（零成本）
                                     └─ 缺失/低置信    → [LLM兜底层] → 只补这些字段

相比分类模块的升级：**字段级兜底**
不是"整封邮件拿不准就全交给 LLM"，而是"哪个字段不行就补哪个"，
规则层已经确认的字段不再重复花钱。

用法：
    python pipeline_extract.py [数据文件] [--threshold 0.7] [--no-llm]

作者：麦当
日期：2026-08-31
"""

import os
import sys
import json
import argparse

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

_PKGS = r"C:\Users\Administrator\.workbuddy\binaries\python\pkgs"
if _PKGS not in sys.path:
    sys.path.insert(0, _PKGS)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from extractor import extract_fields, FIELDS, is_hit
from llm_extract import ExtractManager

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_DATA = os.path.join(BASE_DIR, "data", "extract_emails.json")
DEFAULT_OUT = os.path.join(BASE_DIR, "output", "提取结果.xlsx")

LABEL_CN = {
    "customer": "客户", "product": "产品", "quantity": "数量",
    "price": "价格", "deadline": "截止日期",
}


def process(emails, threshold=0.7, use_llm=True):
    em = ExtractManager(threshold) if use_llm else None
    rows = []

    for mail in emails:
        rule = extract_fields(mail)
        expected = mail.get("expected", {})
        expected = expected if isinstance(expected, dict) else {}
        targets = em.need_fields(rule) if em else []
        llm_vals, llm_note = (em.extract(mail, rule, targets) if em else ({}, "兜底关闭"))

        for f in FIELDS:
            rv = rule[f]["value"]
            conf = rule[f]["confidence"]
            exp = expected.get(f, "")
            rule_hit = is_hit(rv, exp, f)

            if f in targets:
                if llm_vals:
                    lv = llm_vals.get(f, None)
                    fv = lv
                    src = "LLM" if lv is not None else "LLM(确认未提供)"
                else:
                    lv = None
                    fv = rv
                    src = "规则层(降级)"
            else:
                lv = None
                fv = rv
                src = "规则层"

            final_hit = is_hit(fv, exp, f)
            rows.append({
                "id": mail.get("id", ""),
                "field": LABEL_CN.get(f, f),
                "rule_value": rv if rv is not None else "（未抽出）",
                "confidence": round(conf, 2),
                "need_llm": f in targets,
                "llm_value": lv if lv is not None else ("-" if not (f in targets) else "null"),
                "final_value": fv if fv is not None else "（未提供）",
                "source": src,
                "expected": exp if exp else "（未提供）",
                "rule_hit": rule_hit,
                "final_hit": final_hit,
                "note": mail.get("note", ""),
            })

    return rows, em


def export_excel(rows, path, meta):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = openpyxl.Workbook()

    # ---------- Sheet1 字段级明细 ----------
    ws = wb.active
    ws.title = "字段级明细"
    headers = ["邮件ID", "字段", "规则层值", "置信度", "触发兜底", "LLM值",
               "最终值", "来源", "正确答案", "规则层命中", "最终命中", "效果"]
    ws.append(headers)
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in ws[1]:
        c.fill = hf
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    right = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    wrong = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fb = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for r in rows:
        if not r["rule_hit"] and r["final_hit"]:
            effect = "✓ 兜底挽回"
        elif r["rule_hit"] and not r["final_hit"]:
            effect = "✗ 兜底改错"
        elif r["final_hit"]:
            effect = "✓ 命中"
        else:
            effect = "✗ 仍错"

        ws.append([
            r["id"], r["field"], r["rule_value"], f"{r['confidence']:.0%}",
            "是" if r["need_llm"] else "-", r["llm_value"], r["final_value"], r["source"],
            r["expected"], "✓" if r["rule_hit"] else "✗", "✓" if r["final_hit"] else "✗",
            effect,
        ])
        i = ws.max_row
        if r["need_llm"]:
            for c in ws[i]:
                c.fill = fb
        cell = ws.cell(row=i, column=12)
        if effect.startswith("✓ 兜底挽回"):
            cell.fill = right
            cell.font = Font(bold=True, color="006100")
        elif effect.startswith("✗"):
            cell.fill = wrong
            cell.font = Font(bold=True, color="9C0006")

    for i, w in enumerate([8, 10, 30, 9, 10, 26, 30, 16, 30, 11, 11, 13], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---------- Sheet2 汇总 ----------
    ws2 = wb.create_sheet("汇总")
    total = len(rows)
    rule_ok = sum(1 for r in rows if r["rule_hit"])
    final_ok = sum(1 for r in rows if r["final_hit"])
    fb_fields = sum(1 for r in rows if r["need_llm"])
    saved = sum(1 for r in rows if (not r["rule_hit"] and r["final_hit"]))
    broken = sum(1 for r in rows if (r["rule_hit"] and not r["final_hit"]))

    ws2.append(["【信息提取流水线 · 汇总】"])
    ws2.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws2.append([])
    for k, v in [
        ("样本邮件数", len({r["id"] for r in rows})),
        ("字段总数", total),
        ("规则层命中", f"{rule_ok}/{total} = {rule_ok/total:.1%}"),
        ("兜底后命中", f"{final_ok}/{total} = {final_ok/total:.1%}"),
        ("触发兜底字段", f"{fb_fields}/{total} = {fb_fields/total:.1%}"),
        ("兜底挽回字段", saved),
        ("兜底改错字段", broken),
        ("净提升", f"{(final_ok - rule_ok):+d} 个字段"),
        ("LLM Provider", f"{meta['provider']}{'（模拟）' if meta['is_mock'] else '（真实）'}"),
        ("LLM 调用次数（邮件级）", meta["calls"]),
        ("LLM 补全字段数", meta.get("field_calls", 0)),
    ]:
        ws2.append([k, v])
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 34

    wb.save(path)
    return path


def main():
    ap = argparse.ArgumentParser(description="信息提取流水线 v1.0")
    ap.add_argument("data", nargs="?", default=DEFAULT_DATA)
    ap.add_argument("--threshold", type=float, default=0.7)
    ap.add_argument("--no-llm", action="store_true", help="只跑规则层，用于对比")
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()

    with open(args.data, "r", encoding="utf-8") as f:
        emails = json.load(f)

    print("=" * 90)
    print("信息提取流水线 v1.0  ——  规则层 + 字段级 LLM 兜底")
    print("=" * 90)
    print(f"数据文件: {os.path.basename(args.data)}   样本: {len(emails)} 封   阈值: {args.threshold:.0%}")
    print(f"LLM 兜底: {'关闭（仅规则层）' if args.no_llm else '开启'}\n")

    rows, em = process(emails, threshold=args.threshold, use_llm=not args.no_llm)
    meta = em.summary() if em else {"provider": "none", "is_mock": True, "calls": 0, "field_calls": 0}

    if em:
        tag = "（模拟模式）" if meta["is_mock"] else "（真实 API）"
        print(f"LLM Provider: {meta['provider']} {tag}\n")

    cur = None
    print(f"{'ID':<6}{'字段':<8}{'规则层值':<26}{'置信':<7}{'LLM':<6}{'最终值':<30}{'命中':<6}")
    print("-" * 90)
    for r in rows:
        if cur != r["id"]:
            print("-" * 90)
            cur = r["id"]
        mark = "✓" if r["final_hit"] else "✗"
        if not r["rule_hit"] and r["final_hit"]:
            mark = "✓挽回"
        print(f"{r['id']:<6}{r['field']:<8}{str(r['rule_value'])[:24]:<26}"
              f"{r['confidence']:<7.0%}{('是' if r['need_llm'] else '-'):<6}"
              f"{str(r['final_value'])[:28]:<30}{mark:<6}")

    total = len(rows)
    rule_ok = sum(1 for r in rows if r["rule_hit"])
    final_ok = sum(1 for r in rows if r["final_hit"])
    fb_fields = sum(1 for r in rows if r["need_llm"])

    print("-" * 90)
    print(f"规则层字段准确率: {rule_ok}/{total} = {rule_ok/total:.1%}")
    print(f"兜底后字段准确率: {final_ok}/{total} = {final_ok/total:.1%}")
    print(f"触发兜底字段:     {fb_fields}/{total} = {fb_fields/total:.1%}")
    print(f"净提升:           {final_ok - rule_ok:+d} 个字段")
    print(f"LLM 调用:         {meta['calls']} 次（邮件级），覆盖 {meta.get('field_calls', 0)} 个字段")

    out = export_excel(rows, args.out, meta)
    print("=" * 90)
    print(f"[已导出] {out}")
    print("=" * 90)
    return 0


if __name__ == "__main__":
    sys.exit(main())
