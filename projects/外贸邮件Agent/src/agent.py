# -*- coding: utf-8 -*-
"""
外贸邮件智能处理 Agent v1.0 —— 端到端编排
==========================================
流程：

    邮件 ──→ [① 分类] 这是什么邮件？（询盘/订单/投诉/通知）
         ──→ [② 提取] 关键信息是什么？（客户/产品/数量/价格/截止日）
         ──→ [③ 决策] 现在该做什么？（优先级/建议动作/风险提示）
                                          ↓
                                   CRM 记录草稿

能力分层与分工：
    ① 分类：规则层(关键词加权) + LLM 兜底   —— 语义歧义交给 LLM
    ② 提取：规则层(正则)     + LLM 兜底   —— 型号规格、引用历史仲裁交给 LLM
    ③ 决策：纯规则                        —— 见下方说明

为什么决策层不用 LLM：
    企业场景里"为什么把这封标成高优先级"必须能说清楚，规则决策可解释、
    可审计、零成本、毫秒级。而这部分逻辑本身并不复杂，
    LLM 的语义能力应该花在刀刃上——分类和提取这两个真正的难点。

用法：
    python agent.py [数据文件] [--cls-threshold 0.8] [--ext-threshold 0.7]

作者：麦当
日期：2026-08-31
"""

import os
import sys
import re
import json
import argparse
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

_PKGS = r"C:\Users\Administrator\.workbuddy\binaries\python\pkgs"
if _PKGS not in sys.path:
    sys.path.insert(0, _PKGS)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from classifier import score_email, LABEL_CN
from llm_fallback import FallbackManager
from extractor import extract_fields, FIELDS, is_hit
from llm_extract import ExtractManager

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_DATA = os.path.join(BASE_DIR, "data", "e2e_emails.json")
DEFAULT_OUT = os.path.join(BASE_DIR, "output", "Agent处理结果.xlsx")

FIELD_CN = {
    "customer": "客户", "product": "产品", "quantity": "数量",
    "price": "价格", "deadline": "截止日",
}

# 基准日：用于计算"距截止日还剩几天"
BASE_DATE = date.today()

MONTH_MAP = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
             "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}


# ---------------------------------------------------------------------------
# 决策层
# ---------------------------------------------------------------------------
def _diff_days(y, mo, d, base):
    try:
        return (date(y, mo, d) - base).days
    except ValueError:
        return None


def parse_deadline_days(value, base=BASE_DATE):
    """把截止日表述解析为"距基准日天数"；相对表述或解析失败返回 None"""
    if not value:
        return None
    v = str(value).strip()

    m = re.search(r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日", v)
    if m:
        y, mo, d = map(int, m.groups())
        return _diff_days(y, mo, d, base)

    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", v)
    if m:
        y, mo, d = map(int, m.groups())
        return _diff_days(y, mo, d, base)

    m = re.search(r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]{3,9})\s+(\d{4})", v)
    if m:
        mo = MONTH_MAP.get(m.group(2)[:3].lower())
        if mo:
            return _diff_days(int(m.group(3)), mo, int(m.group(1)), base)

    m = re.search(r"([A-Za-z]{3,9})\s+(\d{1,2})(?:st|nd|rd|th)?,?\s+(\d{4})", v)
    if m:
        mo = MONTH_MAP.get(m.group(1)[:3].lower())
        if mo:
            return _diff_days(int(m.group(3)), mo, int(m.group(2)), base)

    return None


def decide(category, fields, mail, base=BASE_DATE):
    """
    决策：给出优先级、判断依据、建议动作、风险提示。
    纯规则实现 —— 每一条结论都能追溯到具体原因（可审计）。
    """
    subject = mail.get("subject", "")
    body = mail.get("body", "")
    text = f"{subject} {body}".lower()

    priority, reasons, actions, risks = "中", [], [], []

    # 1) 紧急表述
    if re.search(r"urgent|asap|immediately|紧急|尽快|务必", text):
        priority = "高"
        reasons.append("含紧急表述（URGENT/ASAP/紧急）")

    # 2) 客诉天然高优先级
    if category == "complaint":
        priority = "高"
        reasons.append("客诉类，涉及索赔与客户关系")

    # 3) 截止日临近程度
    days = parse_deadline_days(fields.get("deadline"), base)
    if days is not None:
        if days < 0:
            priority = "高"
            reasons.append(f"截止日已过期 {abs(days)} 天")
        elif days <= 14:
            priority = "高"
            reasons.append(f"距截止日仅 {days} 天")
        elif days <= 30:
            if priority != "高":
                priority = "中"
            reasons.append(f"距截止日 {days} 天")
    elif fields.get("deadline"):
        reasons.append("截止日为相对表述，无法自动计算剩余天数")

    # 4) 按分类给出建议动作
    action = {
        "inquiry": "准备 PI 报价单：确认单价 / 交期 / 付款方式 / 报价有效期",
        "order": "确认订单：核对 PO 号与数量 → 排产 → 回签",
        "complaint": "启动客诉流程：核实批次 → 判定责任 → 48h 内给出处理方案",
        "notification": "知悉归档：同步物流节点给业务与跟单，更新项目台账",
    }.get(category, "转人工判读")
    actions.append(action)

    # 5) 风险：关键字段缺失（不同类别关注的关键字段不同）
    key_fields = {
        "inquiry": ["quantity", "deadline"],
        "order": ["quantity", "deadline"],
        "complaint": ["quantity", "deadline"],
        "notification": ["quantity"],
    }.get(category, [])
    missing = [f for f in key_fields if not fields.get(f)]
    if missing:
        risks.append("缺关键信息（" + "、".join(FIELD_CN[f] for f in missing)
                     + "）→ 需人工向客户确认后才能推进")

    # 6) 价格缺失对询盘/订单是明显风险
    if category in ("inquiry", "order") and not fields.get("price"):
        risks.append("未提供价格 → 属询价阶段或沿用前合同价，报价前需确认")

    if not reasons:
        reasons.append("无紧急信号，按常规流程处理")

    return priority, reasons, actions, risks


# ---------------------------------------------------------------------------
# Agent
# ---------------------------------------------------------------------------
class MailAgent:
    """端到端：分类 → 提取 → 决策"""

    def __init__(self, cls_threshold=0.8, ext_threshold=0.7):
        self.fm = FallbackManager(cls_threshold)
        self.em = ExtractManager(ext_threshold)

    def process_one(self, mail):
        subject = mail.get("subject", "")
        body = mail.get("body", "")

        # ---- ① 分类 ----
        rule_cat, rule_conf, scores, hits = score_email(subject, body)
        need_cls = self.fm.need_fallback(rule_conf, scores)   # 含争议度检测
        llm_cat = ""
        if need_cls:
            try:
                llm_cat, _, _ = self.fm.classify(subject, body, scores)
            except Exception:
                llm_cat = ""
            final_cat = llm_cat or rule_cat
            cat_source = "LLM" if llm_cat else "规则层(降级)"
        else:
            final_cat, cat_source = rule_cat, "规则层"

        # ---- ② 提取 ----
        rule_fields = extract_fields(mail)
        targets = self.em.need_fields(rule_fields)
        # 分类驱动提取：把分类结果传给提取层，让它知道各字段该取哪个值
        llm_vals, _ = self.em.extract(mail, rule_fields, targets, category=final_cat)

        final_fields = {}
        for f in FIELDS:
            if f in targets and llm_vals:
                final_fields[f] = llm_vals.get(f)
            else:
                final_fields[f] = rule_fields[f]["value"]

        # ---- ③ 决策 ----
        priority, reasons, actions, risks = decide(final_cat, final_fields, mail)

        return {
            "id": mail.get("id", ""),
            "subject": subject,
            "category": final_cat,
            "cat_source": cat_source,
            "need_cls": need_cls,
            "rule_cat": rule_cat,
            "rule_conf": round(rule_conf, 3),
            "expected_cat": mail.get("category", ""),
            "fields": final_fields,
            "rule_fields": {f: rule_fields[f]["value"] for f in FIELDS},
            "ext_targets": targets,
            "priority": priority,
            "reasons": reasons,
            "actions": actions,
            "risks": risks,
            "expected": mail.get("expected", {}),
        }

    def process(self, emails):
        return [self.process_one(m) for m in emails]


# ---------------------------------------------------------------------------
# 输出
# ---------------------------------------------------------------------------
def export_excel(results, path, meta):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = openpyxl.Workbook()

    # ---------- Sheet1 处理结果（一封一行，可直接当 CRM 草稿） ----------
    ws = wb.active
    ws.title = "处理结果"
    headers = ["邮件ID", "主题", "分类", "分类来源", "客户", "产品", "数量",
               "价格", "截止日", "优先级", "判断依据", "建议动作", "风险提示"]
    ws.append(headers)
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in ws[1]:
        c.fill = hf
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    high = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    mid = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    low = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for r in results:
        f = r["fields"]
        ws.append([
            r["id"], r["subject"][:40], LABEL_CN.get(r["category"], r["category"]),
            r["cat_source"],
            f.get("customer") or "—", f.get("product") or "—",
            f.get("quantity") or "—", f.get("price") or "—",
            f.get("deadline") or "—",
            r["priority"],
            "；".join(r["reasons"]),
            "；".join(r["actions"]),
            "；".join(r["risks"]) or "无",
        ])
        i = ws.max_row
        cell = ws.cell(row=i, column=10)
        cell.fill = {"高": high, "中": mid}.get(r["priority"], low)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for i, w in enumerate([8, 38, 8, 12, 22, 30, 26, 24, 22, 8, 34, 40, 36], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---------- Sheet2 准确率 ----------
    n = len(results)
    cls_rule_ok = sum(1 for r in results if r["rule_cat"] == r["expected_cat"])
    cls_final_ok = sum(1 for r in results if r["category"] == r["expected_cat"])

    total_f = n * len(FIELDS)
    ext_rule_ok = sum(1 for r in results for f in FIELDS
                      if is_hit(r["rule_fields"].get(f), r["expected"].get(f, ""), f))
    ext_final_ok = sum(1 for r in results for f in FIELDS
                       if is_hit(r["fields"].get(f), r["expected"].get(f, ""), f))

    ws2 = wb.create_sheet("准确率")
    ws2.append(["【端到端 Agent · 效果汇总】"])
    ws2.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws2.append([])
    ws2.append(["模块", "规则层", "最终（含LLM兜底）", "提升"])
    for c in ws2[3]:
        c.font = Font(bold=True)
    ws2.append(["① 分类", f"{cls_rule_ok}/{n} = {cls_rule_ok/n:.1%}",
                f"{cls_final_ok}/{n} = {cls_final_ok/n:.1%}",
                f"{cls_final_ok - cls_rule_ok:+d} 封"])
    ws2.append(["② 提取（字段级）", f"{ext_rule_ok}/{total_f} = {ext_rule_ok/total_f:.1%}",
                f"{ext_final_ok}/{total_f} = {ext_final_ok/total_f:.1%}",
                f"{ext_final_ok - ext_rule_ok:+d} 个字段"])
    ws2.append([])
    ws2.append(["③ 决策（规则层）", "优先级分布：" +
                "  ".join(f"{p} {sum(1 for r in results if r['priority'] == p)} 封"
                          for p in ("高", "中", "低"))])
    ws2.append([])
    for k, v in [
        ("样本邮件数", n),
        ("分类触发 LLM", f"{sum(1 for r in results if r['need_cls'])} 封"),
        ("提取触发 LLM", f"{sum(len(r['ext_targets']) for r in results)} 个字段 / "
                        f"{sum(1 for r in results if r['ext_targets'])} 封"),
        ("LLM 调用总计", f"{meta['cls_calls'] + meta['ext_calls']} 次"
                        f"（分类 {meta['cls_calls']} + 提取 {meta['ext_calls']}）"),
        ("Provider", f"{meta['provider']}{'（模拟）' if meta['is_mock'] else '（真实）'}"),
    ]:
        ws2.append([k, v])
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 34
    ws2.column_dimensions["C"].width = 30

    wb.save(path)
    return {
        "n": n, "cls_rule_ok": cls_rule_ok, "cls_final_ok": cls_final_ok,
        "total_f": total_f, "ext_rule_ok": ext_rule_ok, "ext_final_ok": ext_final_ok,
        "path": path,
    }


def main():
    ap = argparse.ArgumentParser(description="外贸邮件智能处理 Agent v1.0")
    ap.add_argument("data", nargs="?", default=DEFAULT_DATA)
    ap.add_argument("--cls-threshold", type=float, default=0.8)
    ap.add_argument("--ext-threshold", type=float, default=0.7)
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()

    with open(args.data, "r", encoding="utf-8") as f:
        emails = json.load(f)

    print("=" * 88)
    print("外贸邮件智能处理 Agent v1.0  ——  分类 → 提取 → 决策")
    print("=" * 88)
    print(f"数据文件: {os.path.basename(args.data)}   样本: {len(emails)} 封")
    print(f"基准日期: {BASE_DATE}（用于计算距截止日天数）\n")

    agent = MailAgent(args.cls_threshold, args.ext_threshold)
    print(f"LLM Provider: {agent.em.provider.name} "
          f"{'（模拟）' if agent.em.is_mock else '（真实）'}\n")

    results = agent.process(emails)

    print(f"{'ID':<6}{'分类':<7}{'来源':<9}{'优先级':<7}{'客户':<20}{'数量':<18}{'截止日':<18}")
    print("-" * 88)
    for r in results:
        f = r["fields"]
        print(f"{r['id']:<6}{LABEL_CN.get(r['category'], r['category']):<7}{r['cat_source']:<9}"
              f"{r['priority']:<7}{str(f.get('customer') or '—')[:18]:<20}"
              f"{str(f.get('quantity') or '—')[:16]:<18}"
              f"{str(f.get('deadline') or '—')[:16]:<18}")

    meta = {
        "provider": agent.em.provider.name,
        "is_mock": agent.em.is_mock,
        "cls_calls": agent.fm.call_count,
        "ext_calls": agent.em.call_count,
    }
    stat = export_excel(results, args.out, meta)

    print("-" * 88)
    print(f"① 分类准确率:   规则层 {stat['cls_rule_ok']}/{stat['n']} = {stat['cls_rule_ok']/stat['n']:.1%}"
          f"   →  最终 {stat['cls_final_ok']}/{stat['n']} = {stat['cls_final_ok']/stat['n']:.1%}")
    print(f"② 提取准确率:   规则层 {stat['ext_rule_ok']}/{stat['total_f']} = {stat['ext_rule_ok']/stat['total_f']:.1%}"
          f"   →  最终 {stat['ext_final_ok']}/{stat['total_f']} = {stat['ext_final_ok']/stat['total_f']:.1%}")
    print(f"③ 优先级分布:   " + "  ".join(
        f"{p} {sum(1 for r in results if r['priority'] == p)} 封" for p in ("高", "中", "低")))
    print(f"LLM 调用:       分类 {meta['cls_calls']} 次 + 提取 {meta['ext_calls']} 次"
          f" = {meta['cls_calls'] + meta['ext_calls']} 次")
    print("=" * 88)
    print(f"[已导出] {stat['path']}")
    print("=" * 88)
    return 0


if __name__ == "__main__":
    sys.exit(main())
